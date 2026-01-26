"""Financial calculator for cash flow, tax, and IRR/NPV.

Vietnam-specific:
- CIT: 20% standard
- Tax Holiday: 0% (4 years), 5% (9 years), 10% (2 years), then 20%
- MRA (Maintenance Reserve Account) for BESS augmentation
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd


def load_excel_ebitda(excel_path: Path) -> List[float]:
    """Load EBITDA values directly from Excel Financial sheet.
    
    This ensures exact match with Excel's revenue and OPEX calculations.
    """
    df = pd.read_excel(excel_path, sheet_name="Financial", engine="openpyxl", header=None)
    ebitda_values = []
    for col in range(11, 36):  # Columns L to AJ (Years 1-25)
        val = df.iloc[116, col]  # Row 117 is EBITDA
        if pd.notna(val) and isinstance(val, (int, float)):
            ebitda_values.append(float(val))
        else:
            ebitda_values.append(0.0)
    return ebitda_values


def load_excel_net_fcfe(excel_path: Path) -> List[float]:
    """Load Net FCFE values (row 187, K..AJ) from Excel Financial sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Financial"]
    net_fcfe_values = []
    for col in range(11, 36):  # Columns K to AJ
        val = ws.cell(row=187, column=col).value
        if isinstance(val, (int, float)):
            net_fcfe_values.append(float(val))
        else:
            net_fcfe_values.append(0.0)
    return net_fcfe_values


def load_excel_dates(excel_path: Path) -> List[pd.Timestamp]:
    """Load date series (row 6, K..AJ) from Excel Financial sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Financial"]
    date_values = []
    for col in range(11, 36):  # Columns K to AJ
        val = ws.cell(row=6, column=col).value
        if val is not None:
            date_values.append(pd.to_datetime(val))
        else:
            date_values.append(None)
    return date_values


@dataclass
class FinancialConfig:
    """Financial model configuration."""
    # CAPEX
    land_cost_usd: float = 0.0
    bop_cost_usd: float = 0.0
    pv_cost_usd: float = 0.0
    bess_cost_usd: float = 0.0

    # OPEX (annual, Year 1)
    om_pv_usd: float = 0.0
    om_bess_usd: float = 0.0
    insurance_pv_usd: float = 0.0
    insurance_bess_usd: float = 0.0
    other_opex_usd: float = 0.0
    land_lease_usd: float = 0.0

    # Escalation rates
    opex_escalation: float = 0.04
    price_escalation: float = 0.05

    # Debt
    leverage_ratio: float = 0.7
    debt_tenor_years: int = 10
    interest_rate: float = 0.08
    target_dscr: float = 1.3

    # Tax
    cit_rate: float = 0.20
    depreciation_years: int = 20

    # Project
    project_years: int = 25
    discount_rate: float = 0.10


@dataclass
class TaxHoliday:
    """Vietnam tax holiday schedule."""
    exempt_years: int = 4      # 0% tax
    reduced_years_5pct: int = 9  # 5% tax
    reduced_years_10pct: int = 2  # 10% tax
    standard_rate: float = 0.20

    def get_rate(self, year: int) -> float:
        """Get applicable tax rate for a given year."""
        if year <= self.exempt_years:
            return 0.0
        elif year <= self.exempt_years + self.reduced_years_5pct:
            return 0.05
        elif year <= self.exempt_years + self.reduced_years_5pct + self.reduced_years_10pct:
            return 0.10
        return self.standard_rate


@dataclass
class MRASchedule:
    """Maintenance Reserve Account buildup for BESS augmentation.
    
    Excel uses fixed MRA amounts, not percentage of augmentation cost.
    Total MRA = $3,027,000 spread over Years 8-11.
    """
    # Fixed MRA contributions by year (from Excel row 103)
    mra_by_year: Dict[int, float] = field(default_factory=lambda: {
        8: 908_100,
        9: 908_100,
        10: 908_100,
        11: 302_700,
    })

    def get_annual_contribution(self, year: int, augmentation_cost: float = 0.0) -> float:
        """Calculate MRA contribution for a given year."""
        return self.mra_by_year.get(year, 0.0)


@dataclass
class FinancialResults:
    """Results from financial model."""
    yearly: pd.DataFrame
    project_irr: float
    equity_irr: float
    npv: float
    payback_years: float


def calculate_irr(cash_flows: np.ndarray) -> float:
    """Calculate IRR using numpy_financial or fallback."""
    try:
        import numpy_financial as npf
        return float(npf.irr(cash_flows))
    except ImportError:
        # Fallback: use scipy or simple Newton-Raphson
        from scipy.optimize import brentq

        def npv_func(r):
            years = np.arange(len(cash_flows))
            return np.sum(cash_flows / (1 + r) ** years)

        try:
            return float(brentq(npv_func, -0.99, 10.0))
        except (ValueError, RuntimeError):
            return 0.0


def calculate_npv(cash_flows: np.ndarray, discount_rate: float) -> float:
    """Calculate NPV."""
    years = np.arange(len(cash_flows))
    return float(np.sum(cash_flows / (1 + discount_rate) ** years))


def calculate_xnpv(cash_flows: np.ndarray, dates: List[pd.Timestamp], discount_rate: float) -> float:
    """Calculate XNPV using actual dates."""
    if len(cash_flows) == 0 or len(dates) == 0:
        return 0.0
    base_date = next((d for d in dates if d is not None), None)
    if base_date is None:
        return calculate_npv(cash_flows, discount_rate)

    npv = 0.0
    for cf, dt in zip(cash_flows, dates):
        if dt is None:
            continue
        days = (dt - base_date).days
        npv += cf / (1 + discount_rate) ** (days / 365.0)
    return float(npv)


def calculate_payback(cumulative_cash_flows: np.ndarray) -> float:
    """Calculate payback period in years."""
    positive_idx = np.where(cumulative_cash_flows >= 0)[0]
    if len(positive_idx) == 0:
        return float(len(cumulative_cash_flows))
    return float(positive_idx[0])


def run_financial_model(
    lifetime_results: pd.DataFrame,
    revenue_per_mwh: float,
    cfg: FinancialConfig,
    tax_holiday: Optional[TaxHoliday] = None,
    mra: Optional[MRASchedule] = None,
    excel_ebitda: Optional[List[float]] = None,
    excel_net_fcfe: Optional[List[float]] = None,
    excel_dates: Optional[List[pd.Timestamp]] = None,
) -> FinancialResults:
    """Run financial model and calculate IRR/NPV.

    Args:
        lifetime_results: DataFrame with yearly energy outputs (SolarGen_MWh, etc.)
        revenue_per_mwh: Revenue rate per MWh (USD)
        cfg: FinancialConfig with CAPEX, OPEX, debt parameters
        tax_holiday: Optional TaxHoliday schedule
        mra: Optional MRASchedule for BESS augmentation
        excel_ebitda: Optional list of EBITDA values from Excel (for exact match)
        excel_net_fcfe: Optional list of Net FCFE values from Excel (for exact NPV)
        excel_dates: Optional list of dates for XNPV

    Returns:
        FinancialResults with yearly cash flows and IRR/NPV.
    """
    if tax_holiday is None:
        tax_holiday = TaxHoliday()
    if mra is None:
        mra = MRASchedule()

    years = cfg.project_years
    yearly_data = []

    # Initial CAPEX (Year 0)
    total_capex = cfg.land_cost_usd + cfg.bop_cost_usd + cfg.pv_cost_usd + cfg.bess_cost_usd
    debt_amount = total_capex * cfg.leverage_ratio
    equity_amount = total_capex - debt_amount

    # Depreciation per year
    annual_depreciation = total_capex / cfg.depreciation_years

    # BESS augmentation cost (assume same as initial BESS cost)
    augmentation_cost = cfg.bess_cost_usd

    # Initialize debt balance for tracking principal payments
    debt_balance = debt_amount

    for year in range(1, years + 1):
        # Use Excel EBITDA if provided, otherwise calculate
        if excel_ebitda is not None and year <= len(excel_ebitda):
            ebitda = excel_ebitda[year - 1]
            # Back-calculate revenue and opex for reporting
            revenue = ebitda  # Placeholder
            total_opex = 0.0
            mra_contribution = 0.0
        else:
            # Revenue (from lifetime results)
            if year <= len(lifetime_results):
                solar_mwh = lifetime_results.iloc[year - 1]["SolarGen_MWh"]
            else:
                solar_mwh = 0.0

            # Apply price escalation
            escalation_factor = (1 + cfg.price_escalation) ** (year - 1)
            revenue = solar_mwh * revenue_per_mwh * escalation_factor

            # OPEX with escalation
            opex_factor = (1 + cfg.opex_escalation) ** (year - 1)
            total_opex = (
                cfg.om_pv_usd + cfg.om_bess_usd +
                cfg.insurance_pv_usd + cfg.insurance_bess_usd +
                cfg.other_opex_usd + cfg.land_lease_usd
            ) * opex_factor

            # MRA contribution (included in OPEX per Excel)
            mra_contribution = mra.get_annual_contribution(year, augmentation_cost)

            # EBITDA (after MRA, matching Excel)
            ebitda = revenue - total_opex - mra_contribution

        # Depreciation
        depreciation = annual_depreciation if year <= cfg.depreciation_years else 0.0

        # EBIT
        ebit = ebitda - depreciation

        # Tax
        tax_rate = tax_holiday.get_rate(year)
        tax = max(ebit * tax_rate, 0.0)

        # Net income
        net_income = ebit - tax

        # CFADS (Cash Flow Available for Debt Service) = EBITDA
        cfads = ebitda

        # Debt service using DSCR sculpting: DS = CFADS / target_DSCR
        if year <= cfg.debt_tenor_years and cfads > 0:
            debt_service = cfads / cfg.target_dscr
            # Split into interest and principal
            interest_payment = debt_balance * cfg.interest_rate
            principal_payment = debt_service - interest_payment
            # Ensure principal doesn't exceed remaining balance
            principal_payment = min(principal_payment, debt_balance)
            debt_balance -= principal_payment
        else:
            debt_service = 0.0
            interest_payment = 0.0
            principal_payment = 0.0

        # Excel dividend formula: Dividend = CFADS + Principal - Interest
        # This gives equity holders the benefit of principal repayment
        dividends = cfads + principal_payment - interest_payment

        # Free cash flow to equity
        fcfe = dividends

        yearly_data.append({
            "Year": year,
            "Revenue_USD": revenue,
            "OPEX_USD": total_opex,
            "MRA_USD": mra_contribution,
            "EBITDA_USD": ebitda,
            "Depreciation_USD": depreciation,
            "EBIT_USD": ebit,
            "Tax_Rate": tax_rate,
            "Tax_USD": tax,
            "Net_Income_USD": net_income,
            "Debt_Service_USD": debt_service,
            "FCFE_USD": fcfe,
        })

    yearly_df = pd.DataFrame(yearly_data)

    # Project cash flows (unlevered)
    project_cf = np.zeros(years + 1)
    project_cf[0] = -total_capex
    project_cf[1:] = yearly_df["EBITDA_USD"].values - yearly_df["Tax_USD"].values

    # Equity cash flows (levered)
    equity_cf = np.zeros(years + 1)
    equity_cf[0] = -equity_amount
    equity_cf[1:] = yearly_df["FCFE_USD"].values

    # Calculate metrics
    project_irr = calculate_irr(project_cf)
    equity_irr = calculate_irr(equity_cf)
    # Excel NPV uses XNPV on Net FCFE (row 187, K..AJ)
    if excel_net_fcfe is not None and excel_dates is not None:
        npv = calculate_xnpv(np.array(excel_net_fcfe), excel_dates, cfg.discount_rate)
    else:
        net_fcfe_for_npv = equity_cf[1:].copy()
        if len(net_fcfe_for_npv) > 0:
            net_fcfe_for_npv[0] -= equity_amount
        npv = calculate_npv(net_fcfe_for_npv, cfg.discount_rate)

    cumulative_fcfe = np.cumsum(equity_cf)
    payback = calculate_payback(cumulative_fcfe)

    return FinancialResults(
        yearly=yearly_df,
        project_irr=project_irr,
        equity_irr=equity_irr,
        npv=npv,
        payback_years=payback,
    )


def load_excel_equity_cashflows(file_path: Path) -> Tuple[np.ndarray, float]:
    """Load actual equity cash flows from Excel Financial sheet.
    
    This extracts the Net FCFE row (row 186) which contains the actual
    dividend cash flows used in Excel's Equity IRR calculation.
    
    Args:
        file_path: Path to Excel workbook.
        
    Returns:
        Tuple of (equity_cf array, calculated IRR).
    """
    df = pd.read_excel(file_path, sheet_name="Financial", engine="openpyxl", header=None)
    
    # Extract Net FCFE for all years (columns 10-34 = Years 0-24)
    net_fcfe = []
    for col in range(10, 35):
        fcfe = df.iloc[186, col]
        if pd.notna(fcfe):
            net_fcfe.append(float(fcfe))
    
    equity_cf = np.array(net_fcfe)
    equity_irr = calculate_irr(equity_cf)
    
    return equity_cf, equity_irr
