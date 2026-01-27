"""Scenario runner for synthetic Excel/Python validation.

Creates scenario-specific Excel copies, applies input overrides, recalculates Excel
(if available), then runs the audit comparison against the Python model.
"""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List

from excel_replica.outputs.audit_report import run_audit


@dataclass
class ScenarioOverride:
    sheet: str
    cell: str
    value: Any


@dataclass
class Scenario:
    name: str
    description: str
    overrides: List[ScenarioOverride]


def _load_scenario(file_path: Path) -> Scenario:
    data = json.loads(file_path.read_text(encoding="utf-8"))
    overrides = [ScenarioOverride(**item) for item in data.get("overrides", [])]
    return Scenario(
        name=data.get("name", file_path.stem),
        description=data.get("description", ""),
        overrides=overrides,
    )


def _apply_overrides(excel_path: Path, overrides: List[ScenarioOverride]) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=False)
    for override in overrides:
        ws = wb[override.sheet]
        ws[override.cell].value = override.value
    wb.save(excel_path)


def _recalculate_with_excel(excel_path: Path) -> bool:
    """Attempt to force Excel recalc via COM automation (Windows only).

    Returns True if recalc succeeded, False if COM not available.
    """
    try:
        import win32com.client as win32
    except Exception:
        return False

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(excel_path))
        wb.RefreshAll()
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
        return True
    finally:
        excel.Quit()


def _run_scenario(base_excel: Path, scenario: Scenario, output_dir: Path) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    scenario_excel = output_dir / f"{scenario.name}.xlsx"
    shutil.copy(base_excel, scenario_excel)

    _apply_overrides(scenario_excel, scenario.overrides)
    recalculated = _recalculate_with_excel(scenario_excel)

    comparisons, _ = run_audit(scenario_excel)
    return {
        "scenario": scenario.name,
        "description": scenario.description,
        "excel_path": str(scenario_excel),
        "recalculated": recalculated,
        "comparisons": comparisons,
    }


def _summarize(results: List[Dict[str, Any]]) -> None:
    print("\n=== Scenario Summary ===")
    for result in results:
        scenario = result["scenario"]
        recalculated = result["recalculated"]
        status = "recalc" if recalculated else "manual"
        max_err = max(comp.percent_error for comp in result["comparisons"]) * 100
        print(f"- {scenario}: max error {max_err:.2f}% ({status})")


def main() -> None:
    parser = argparse.ArgumentParser(description="Run synthetic scenario validation")
    parser.add_argument("--base-excel", type=Path, required=True)
    parser.add_argument("--scenarios-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    scenario_files = sorted(args.scenarios_dir.glob("*.json"))
    if not scenario_files:
        raise SystemExit("No scenario JSON files found.")

    results = []
    for scenario_file in scenario_files:
        scenario = _load_scenario(scenario_file)
        result = _run_scenario(args.base_excel, scenario, args.output_dir)
        results.append(result)

    _summarize(results)
    print("\nNote: If recalculation is 'manual', open the scenario Excel file once in Excel, let it recalc, save, then rerun.")


if __name__ == "__main__":
    main()
